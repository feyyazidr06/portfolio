from django.shortcuts import render
from django.urls import reverse
from django.http import HttpResponseRedirect
from .models import Musteri, Note, Company, Doc
import openpyxl, datetime

def exc():
    wb = openpyxl.load_workbook(r"C:\Users\feyya\Downloads\musteri.xlsx")
    # Musteri
    ws = wb.active
    for i in range(2, 4):
        xldate = ws["A" + str(i)].value
        name = ws["B" + str(i)].value
        tel = str(ws["C" + str(i)].value)
        brand = ws["D" + str(i)].value
        reason = ws["E" + str(i)].value
        solution = ws["F" + str(i)].value
        price = str(ws["G" + str(i)].value)

        date_object = datetime.datetime.strptime("-".join(xldate.split(".", 3)), '%d-%m-%Y').date()

        m = Musteri(date=date_object,
            name=name,
            tel=tel,
            brand=brand,
            reason=reason,
            solution=solution,
            price=price,
            active=False)
        m.save()

# Create your views here.
def index(request):
    return render(request, "customer/index.html", {
        "items": Musteri.objects.filter(active=True),
    })

def notes(request):
    return render(request, "customer/notes.html", {
        "notes": Note.objects.all()
    })

def deletenote(request, id):
    Note.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('notes'))

def editnote(request, id):
    if request.method == "POST":
        n = Note.objects.get(pk=id)
        n.title = request.POST["title"]
        n.content = request.POST["content"]
        n.save()
        return HttpResponseRedirect(reverse('notes'))
    return render(request, "customer/note.html", {
        "title": Note.objects.get(pk=id).title,
        "content": Note.objects.get(pk=id).content,
        "edit": True,
    })

def addnote(request):
    n = Note()
    n.save()
    return HttpResponseRedirect(reverse('editnote', args=(n.id,)))

def note(request, id):
    return render(request, "customer/note.html", {
        "id": id,
        "title": Note.objects.get(pk=id).title,
        "content": Note.objects.get(pk=id).content,
        "edit": False,
    })

def liste(request):
    return render(request, "customer/list.html", {
        "items": list(reversed(Musteri.objects.all()))
    })

def search(request):
    if request.method == "POST":
        searched = request.POST['q'].upper()
        arr = Musteri.objects.filter(id__contains=searched) | Musteri.objects.filter(name__contains=searched) | Musteri.objects.filter(tel__contains=searched) | Musteri.objects.filter(brand__contains=searched) | Musteri.objects.filter(model__contains=searched)
        return render(request, "customer/list.html", {
            "items": arr,
        })
    return HttpResponseRedirect(reverse('list'))

def customer(request, id):
    c = Musteri.objects.get(pk=id)
    if request.method == "POST":
        c.date = request.POST['date']
        c.name = request.POST['name'].upper()
        c.tel = request.POST['phone']
        c.brand = request.POST['brand'].upper()
        c.model = request.POST['model'].upper()
        c.reason = request.POST['reason'].upper()
        c.solution = request.POST['solution'].upper()
        c.price = request.POST['price']
        if request.POST['fin-date']:
            c.fin_date = request.POST['fin-date']

        if "active" in request.POST:
            if request.POST['active'] == "on":
                c.active = True
            else:
                c.active = request.POST.get("active", False)
        else:
            c.active = False

        if "foot" in request.POST:
            if request.POST['foot'] == "on":
                c.foot = True
            else:
                c.foot = request.POST.get("foot", False)
        else:
            c.foot = False

        if "drzac" in request.POST:
            if request.POST['drzac'] == "on":
                c.drzac = True
            else:
                c.drzac = request.POST.get("drzac", False)
        else:
            c.drzac = False

        if "cable" in request.POST:
            if request.POST['cable'] == "on":
                c.cabel = True
            else:
                c.cabel = request.POST.get("cable", False)
        else:
            c.cabel = False

        if "module" in request.POST:
            if request.POST['module'] == "on":
                c.module = True
            else:
                c.module = request.POST.get("module", False)
        else:
            c.module = False

        if "remote" in request.POST:
            if request.POST['remote'] == "on":
                c.remote = True
            else:
                c.remote = request.POST.get("remote", False)
        else:
            c.remote = False

        c.save()
        return render(request, "customer/customer.html", {
        "customer": c,
        "success": True,
        })
    return render(request, "customer/customer.html", {
        "customer": c
    })

def add(request):
    today = datetime.date.today()
    c = Musteri(date=today)
    c.save()
    return HttpResponseRedirect(reverse('customer', args=(c.id,)))

def delete(request, id):
    Musteri.objects.get(pk=id).delete()
    return HttpResponseRedirect(reverse('list'))

def finished(request, id):
    today = datetime.date.today()
    c = Musteri.objects.get(pk=id)
    c.active = False
    c.fin_date = today
    c.save()
    return HttpResponseRedirect(reverse('customer', args=(id,)))

def company(request):
    if request.method == "POST":
        c = Company.objects.get()
        c.name = request.POST["company-name"]
        c.location = request.POST["company-loc"]
        c.phone = request.POST["company-phone"]
        if len(request.FILES) != 0:
            c.logo = request.FILES["img"]
        c.save()
        return render(request, "customer/company.html", {
            "company": Company.objects.get(),
            "success": True,
        })
    return render(request, "customer/company.html", {
        "company": Company.objects.get()
    })

def doc(request):
    if request.method == "POST":
        d = Doc.objects.get()
        d.content = request.POST["content"]
        d.save()
        return render(request, "customer/doc.html", {
            "content": Doc.objects.get().content,
            "success": True,
        })
    return render(request, "customer/doc.html", {
        "content": Doc.objects.get().content
    })

def warranty(request, id):
    return render(request, "customer/warranty.html", {
        "company": Company.objects.get(),
        "c": Musteri.objects.get(pk=id),
        "content": Doc.objects.get().content,
    })